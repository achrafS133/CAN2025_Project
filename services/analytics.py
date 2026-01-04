"""
Advanced Analytics Dashboard
Historical analysis, Excel export, anomaly detection, predictive insights
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Tuple
from datetime import datetime, timedelta
from pathlib import Path
import plotly.graph_objects as go
import plotly.express as px
from sklearn.ensemble import IsolationForest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference

from core.logger import app_logger


class AnalyticsEngine:
    """Advanced analytics with ML-based insights"""

    def __init__(self):
        self.data_file = Path("./data/incidents_history.csv")
        self.data_file.parent.mkdir(exist_ok=True)

    def load_historical_data(self, days: int = 30) -> pd.DataFrame:
        """Load historical incident data"""
        if not self.data_file.exists():
            return pd.DataFrame()

        df = pd.DataFrame(
            columns=["timestamp", "type", "location", "severity", "confidence"]
        )
        df["timestamp"] = pd.to_datetime(df["timestamp"])

        # Filter by date range
        cutoff = datetime.now() - timedelta(days=days)
        df = df[df["timestamp"] >= cutoff]

        return df

    def detect_anomalies(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        ML-based anomaly detection using Isolation Forest

        Returns:
            DataFrame with anomaly scores and flags
        """
        if len(df) < 10:
            df["anomaly"] = 0
            df["anomaly_score"] = 0
            return df

        # Feature engineering
        df["hour"] = df["timestamp"].dt.hour
        df["day_of_week"] = df["timestamp"].dt.dayofweek

        # Prepare features
        features = pd.get_dummies(df[["hour", "day_of_week", "type", "location"]])

        # Train Isolation Forest
        model = IsolationForest(contamination=0.1, random_state=42)
        df["anomaly"] = model.fit_predict(features)
        df["anomaly_score"] = model.score_samples(features)

        app_logger.info(
            "anomaly_detection_complete",
            total_incidents=len(df),
            anomalies_detected=sum(df["anomaly"] == -1),
        )

        return df

    def generate_heatmap(self, df: pd.DataFrame) -> go.Figure:
        """Generate incident heatmap by hour and day"""
        if df.empty:
            return go.Figure()

        df["hour"] = df["timestamp"].dt.hour
        df["day"] = df["timestamp"].dt.day_name()

        # Pivot table for heatmap
        pivot = df.groupby(["day", "hour"]).size().unstack(fill_value=0)

        # Order days
        day_order = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ]
        pivot = pivot.reindex(day_order)

        fig = go.Figure(
            data=go.Heatmap(
                z=pivot.values,
                x=pivot.columns,
                y=pivot.index,
                colorscale="Reds",
                text=pivot.values,
                texttemplate="%{text}",
                textfont={"size": 10},
            )
        )

        fig.update_layout(
            title="Incident Heatmap by Day and Hour",
            xaxis_title="Hour of Day",
            yaxis_title="Day of Week",
            height=400,
        )

        return fig

    def generate_trend_analysis(self, df: pd.DataFrame) -> go.Figure:
        """Generate trend analysis with predictions"""
        if df.empty:
            return go.Figure()

        # Daily incident count
        daily = df.groupby(df["timestamp"].dt.date).size().reset_index()
        daily.columns = ["date", "count"]
        daily["date"] = pd.to_datetime(daily["date"])

        # Simple moving average
        daily["MA_7"] = daily["count"].rolling(window=7, min_periods=1).mean()

        fig = go.Figure()

        fig.add_trace(
            go.Scatter(
                x=daily["date"],
                y=daily["count"],
                mode="lines+markers",
                name="Daily Incidents",
                line=dict(color="red", width=2),
            )
        )

        fig.add_trace(
            go.Scatter(
                x=daily["date"],
                y=daily["MA_7"],
                mode="lines",
                name="7-Day Moving Average",
                line=dict(color="blue", width=3, dash="dash"),
            )
        )

        fig.update_layout(
            title="Incident Trends Over Time",
            xaxis_title="Date",
            yaxis_title="Number of Incidents",
            height=400,
            hovermode="x unified",
        )

        return fig

    def comparative_analysis(
        self, df: pd.DataFrame, group_by: str = "location"
    ) -> go.Figure:
        """Compare incidents across locations or types"""
        if df.empty:
            return go.Figure()

        comparison = df.groupby(group_by).size().reset_index()
        comparison.columns = [group_by, "count"]
        comparison = comparison.sort_values("count", ascending=False)

        fig = px.bar(
            comparison,
            x=group_by,
            y="count",
            title=f"Incidents by {group_by.title()}",
            color="count",
            color_continuous_scale="Reds",
        )

        fig.update_layout(height=400)

        return fig

    def export_to_excel(
        self, df: pd.DataFrame, filename: str = "security_analytics_report.xlsx"
    ) -> str:
        """
        Export analytics to Excel with charts and formatting

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Incident Analysis"

        # Header styling
        header_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        # Write headers
        headers = ["Timestamp", "Type", "Location", "Severity", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row[:5], 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=str(value))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary["A1"] = "Total Incidents"
        ws_summary["B1"] = len(df)

        if not df.empty:
            ws_summary["A2"] = "Most Common Type"
            ws_summary["B2"] = (
                df["type"].mode()[0] if len(df["type"].mode()) > 0 else "N/A"
            )

            ws_summary["A3"] = "High Severity Count"
            ws_summary["B3"] = len(df[df["severity"] == "high"])

        # Save file
        output_path = Path(f"./reports/{filename}")
        output_path.parent.mkdir(exist_ok=True)
        wb.save(output_path)

        app_logger.info("excel_report_generated", path=str(output_path))

        return str(output_path)

    def get_predictive_insights(self, df: pd.DataFrame) -> Dict:
        """Generate predictive insights"""
        if df.empty or len(df) < 7:
            return {"prediction": "Insufficient data for predictions", "confidence": 0}

        # Calculate daily averages
        daily_avg = df.groupby(df["timestamp"].dt.date).size().mean()
        recent_7_days = (
            df[df["timestamp"] >= datetime.now() - timedelta(days=7)]
            .groupby(df["timestamp"].dt.date)
            .size()
            .mean()
        )

        # Trend detection
        trend = "increasing" if recent_7_days > daily_avg else "decreasing"
        change_percent = (
            abs((recent_7_days - daily_avg) / daily_avg * 100) if daily_avg > 0 else 0
        )

        return {
            "prediction": f"Incidents are {trend} by {change_percent:.1f}% compared to average",
            "daily_average": round(daily_avg, 1),
            "recent_average": round(recent_7_days, 1),
            "trend": trend,
            "confidence": min(95, 60 + len(df) / 10),  # Confidence increases with data
        }


# Global analytics engine
analytics = AnalyticsEngine()
